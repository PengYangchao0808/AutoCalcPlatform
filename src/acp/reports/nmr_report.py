# pyright: reportMissingTypeStubs=false, reportAny=false, reportExplicitAny=false, reportUnknownVariableType=false, reportOptionalMemberAccess=false, reportUnusedCallResult=false
"""NMR report serialization helpers."""

from __future__ import annotations

import json
import logging
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from acp.nmr.models import NMRReport

logger = logging.getLogger(__name__)


def _convert_paths(value: Any) -> Any:
    """Recursively convert path-like values into strings."""
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {key: _convert_paths(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_convert_paths(item) for item in value]
    if isinstance(value, tuple):
        return [_convert_paths(item) for item in value]
    return value


def _to_cell_value(value: Any) -> Any:
    """Convert nested values into worksheet-safe cell payloads."""
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return json.dumps(_convert_paths(value), ensure_ascii=False)


def write_json_report(report: NMRReport, path: Path) -> None:
    """Serialize an :class:`NMRReport` to JSON."""
    payload = _convert_paths(asdict(report))
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    _ = destination.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def write_xlsx_report(report: NMRReport, path: Path) -> None:
    """Write an :class:`NMRReport` to an Excel workbook."""
    workbook = Workbook()

    summary_sheet = workbook.active
    if summary_sheet is None:
        raise RuntimeError("Workbook did not create a default worksheet")
    summary_sheet.title = "summary"
    summary_sheet.append(["field", "value"])
    summary_rows = [
        ("molecule_name", report.molecule_name),
        ("backend", report.backend),
        ("method", report.method),
        ("basis", report.basis),
        ("temperature_k", report.temperature_k),
        ("n_conformers", len(report.conformers)),
        ("n_averaged_atoms", len(report.averaged_atoms)),
    ]
    for key, value in summary_rows:
        summary_sheet.append([key, _to_cell_value(value)])
    for nucleus, value in sorted(report.references.items()):
        summary_sheet.append([f"reference_{nucleus}", _to_cell_value(value)])
    for key, value in sorted(report.metadata.items()):
        summary_sheet.append([f"metadata_{key}", _to_cell_value(value)])

    averaged_sheet = workbook.create_sheet("averaged_shifts")
    averaged_sheet.append(
        [
            "atom_index",
            "symbol",
            "nucleus",
            "averaged_shielding_ppm",
            "reference_ppm",
            "averaged_shift_ppm",
        ]
    )
    for atom in report.averaged_atoms:
        averaged_sheet.append(
            [
                atom.atom_index,
                atom.symbol,
                atom.nucleus,
                atom.averaged_shielding_ppm,
                atom.reference_ppm,
                atom.averaged_shift_ppm,
            ]
        )

    conformer_sheet = workbook.create_sheet("per_conformer_shifts")
    conformer_sheet.append(
        [
            "record_id",
            "weight",
            "energy_hartree",
            "free_energy_hartree",
            "log_file",
            "atom_index",
            "symbol",
            "nucleus",
            "shielding_ppm",
            "reference_ppm",
            "shift_ppm",
            "anisotropy_ppm",
        ]
    )
    for conformer in report.conformers:
        for shift in conformer.shifts:
            conformer_sheet.append(
                [
                    conformer.record_id,
                    conformer.weight,
                    conformer.energy_hartree,
                    conformer.free_energy_hartree,
                    str(conformer.log_file),
                    shift.atom_index,
                    shift.symbol,
                    shift.nucleus,
                    shift.shielding_ppm,
                    shift.reference_ppm,
                    shift.shift_ppm,
                    shift.anisotropy_ppm,
                ]
            )

    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


__all__ = ["write_json_report", "write_xlsx_report"]
