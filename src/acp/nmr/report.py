# pyright: reportMissingTypeStubs=false, reportExplicitAny=false, reportAny=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownArgumentType=false
"""NMR report serialization (DevDoc §7 / §5 stage 8).

Emits:

* ``nmr_report.json`` — the full machine-readable report (candidates,
  DP4/DP5, assignment tables, regression, per-conformer weights);
* ``nmr_assignment.xlsx`` — per-candidate shift-comparison sheet;
* ``scatter_<nucleus>.png`` / ``error_hist.png`` — diagnostic plots.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

from acp.nmr.models import NmrReport

logger = logging.getLogger(__name__)


def write_json_report(report: NmrReport, output_path: Path) -> Path:
    """Write ``nmr_report.json``."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = report.as_dict()
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return output_path


def write_xlsx_report(report: NmrReport, output_path: Path) -> Path | None:
    """Write ``nmr_assignment.xlsx`` (one sheet per candidate).

    Returns ``None`` (and logs) when openpyxl is unavailable.
    """
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover - openpyxl is in deps
        logger.warning("openpyxl not available; skipping XLSX report")
        return None

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # remove the default sheet — we add one per candidate
    default_ws = wb.active

    for candidate in report.candidates:
        sheet_name = f"cand_{candidate.index}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["atom", "element", "exp_ppm", "calc_ppm", "scaled_ppm", "residual"])
        for assignment in candidate.assignments:
            ws.append(
                [
                    assignment.atom_label,
                    assignment.element,
                    round(assignment.exp_ppm, 4),
                    round(assignment.calc_ppm, 4),
                    round(assignment.scaled_ppm, 4),
                    round(assignment.residual, 4),
                ]
            )
        ws.append([])
        ws.append(["DP4", candidate.dp4_probability])
        ws.append(["DP5", candidate.dp5_probability])
        for nucleus, regression in candidate.regressions.items():
            ws.append([])
            ws.append([f"regression[{nucleus}]", "slope", regression.slope])
            ws.append(["", "intercept", regression.intercept])
            ws.append(["", "r_squared", regression.r_squared])
            ws.append(["", "mae", regression.mae])

    if default_ws is not None and len(wb.sheetnames) > 1:
        wb.remove(default_ws)
    wb.save(output_path)
    return output_path


def write_plots(report: NmrReport, output_dir: Path) -> list[Path]:
    """Write scatter + error-histogram PNGs.

    Returns the list of paths actually written (empty if matplotlib is
    unavailable or the report has no residuals).
    """
    try:
        import matplotlib

        matplotlib.use("Agg")  # headless backend
        import matplotlib.pyplot as plt
    except ImportError:  # pragma: no cover - matplotlib is in deps
        logger.warning("matplotlib not available; skipping plots")
        return []

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    # collect per-nucleus (calc, exp) pairs across candidates
    by_nucleus: dict[str, list[tuple[float, float, int]]] = {}
    for candidate in report.candidates:
        for assignment in candidate.assignments:
            nucleus = _nucleus_of_element(assignment.element)
            by_nucleus.setdefault(nucleus, []).append(
                (assignment.calc_ppm, assignment.exp_ppm, candidate.index)
            )

    for nucleus, triples in by_nucleus.items():
        fig, ax = plt.subplots(figsize=(5, 5))
        calc = [t[0] for t in triples]
        exp = [t[1] for t in triples]
        cand = [t[2] for t in triples]
        scatter = ax.scatter(calc, exp, c=cand, cmap="tab10", alpha=0.7)
        if calc and exp:
            lo = min(min(calc), min(exp))
            hi = max(max(calc), max(exp))
            ax.plot([lo, hi], [lo, hi], "k--", lw=0.8, alpha=0.5)
        ax.set_xlabel(f"calc δ ({nucleus}) / ppm")
        ax.set_ylabel(f"exp δ ({nucleus}) / ppm")
        ax.set_title(f"{nucleus}: calc vs exp")
        fig.colorbar(scatter, ax=ax, label="candidate #")
        fig.tight_layout()
        path = output_dir / f"scatter_{nucleus}.png"
        fig.savefig(path, dpi=120)
        plt.close(fig)
        written.append(path)

    # residual histogram (all nuclei combined, absolute residual)
    all_residuals = [
        assignment.residual
        for candidate in report.candidates
        for assignment in candidate.assignments
    ]
    if all_residuals:
        fig, ax = plt.subplots(figsize=(5, 4))
        ax.hist(all_residuals, bins=20, alpha=0.75, edgecolor="black")
        ax.set_xlabel("residual (δ_exp − δ_scaled) / ppm")
        ax.set_ylabel("count")
        ax.set_title("Residual distribution")
        fig.tight_layout()
        path = output_dir / "error_hist.png"
        fig.savefig(path, dpi=120)
        plt.close(fig)
        written.append(path)

    return written


def write_all_reports(report: NmrReport, output_dir: Path) -> dict[str, Path | None]:
    """Write JSON + XLSX + plots into *output_dir*.

    Returns ``{"json": path, "xlsx": path|None, "plots": [paths...]}``.
    """
    output_dir = Path(output_dir)
    json_path = write_json_report(report, output_dir / "nmr_report.json")
    xlsx_path = write_xlsx_report(report, output_dir / "nmr_assignment.xlsx")
    plots = write_plots(report, output_dir / "plots")
    return {"json": json_path, "xlsx": xlsx_path, "plots": plots}


def _nucleus_of_element(element: str) -> str:
    sym = (element or "").strip()
    if not sym:
        return "?"
    sym = sym[:1].upper() + sym[1:].lower()
    defaults = {"H": "1H", "C": "13C", "N": "15N", "F": "19F", "P": "31P"}
    return defaults.get(sym, f"1{sym}")


__all__ = [
    "write_json_report",
    "write_xlsx_report",
    "write_plots",
    "write_all_reports",
]
