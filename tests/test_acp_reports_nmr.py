# pyright: reportMissingTypeStubs=false, reportAny=false
"""Tests for ACP NMR report writers."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from acp.nmr.models import (
    NMRAveragedAtomResult,
    NMRAtomShift,
    NMRConformerResult,
    NMRReport,
)
from acp.reports import write_json_report, write_xlsx_report


def _make_report(log_file: Path) -> NMRReport:
    return NMRReport(
        molecule_name="ethanol",
        backend="gaussian",
        method="B3LYP",
        basis="def2-TZVPP",
        temperature_k=298.15,
        references={"1H": 31.88, "13C": 186.10},
        conformers=[
            NMRConformerResult(
                record_id="conf_000",
                energy_hartree=-100.0,
                free_energy_hartree=-100.1,
                weight=0.75,
                log_file=log_file,
                shifts=[
                    NMRAtomShift(
                        atom_index=1,
                        symbol="C",
                        nucleus="13C",
                        shielding_ppm=180.0,
                        reference_ppm=186.10,
                        shift_ppm=6.10,
                    )
                ],
            )
        ],
        averaged_atoms=[
            NMRAveragedAtomResult(
                atom_index=1,
                symbol="C",
                nucleus="13C",
                averaged_shielding_ppm=178.5,
                reference_ppm=186.10,
                averaged_shift_ppm=7.6,
            )
        ],
        metadata={"source": "unit-test"},
    )


def test_write_json_report_serializes_nested_paths(tmp_path: Path) -> None:
    log_file = tmp_path / "conf_000.log"
    report = _make_report(log_file)
    json_path = tmp_path / "report.json"

    write_json_report(report, json_path)

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["molecule_name"] == "ethanol"
    assert payload["backend"] == "gaussian"
    assert payload["conformers"][0]["log_file"] == str(log_file)
    assert payload["averaged_atoms"][0]["averaged_shift_ppm"] == 7.6


def test_write_xlsx_report_creates_expected_sheets(tmp_path: Path) -> None:
    report = _make_report(tmp_path / "conf_000.log")
    xlsx_path = tmp_path / "report.xlsx"

    write_xlsx_report(report, xlsx_path)

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["summary", "averaged_shifts", "per_conformer_shifts"]
    assert workbook["summary"]["A2"].value == "molecule_name"
    assert workbook["summary"]["B2"].value == "ethanol"
    assert workbook["averaged_shifts"]["A2"].value == 1
    assert workbook["per_conformer_shifts"]["A2"].value == "conf_000"
